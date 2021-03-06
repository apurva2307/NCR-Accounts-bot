from openpyxl import load_workbook


def make_excel(month, pu, data, lastYearData):
    wb = load_workbook("customFile.xlsx")
    if month[:3] in ["JAN", "FEB", "MAR"]:
        lastYear = int(month[3:]) - 1
    else:
        lastYear = int(month[3:])
    customSheet = wb["Sheet1"]
    customSheet.cell(1, 4).value = f"{pu}"
    customSheet.cell(1, 10).value = "Fig in crore"
    customSheet.cell(3, 2).value = f"20{lastYear-1}-{lastYear}"
    customSheet.cell(3, 3).value = f"20{lastYear}-{lastYear+1}"
    customSheet.cell(3, 4).value = f"{month[:3]}' {int(month[3:])-1}"
    customSheet.cell(3, 5).value = f"{month[:3]}' {month[3:]}"
    customSheet.cell(3, 6).value = f"{month[:3]}' {month[3:]}"
    budget = data[pu]["budget"]
    toEndActualsCoppy = data[pu]["toEndActualsCoppy"]
    toEndBp = data[pu]["toEndBp"]
    toEndActuals = data[pu]["toEndActuals"]
    lastFullYearActuals = lastYearData[pu]["toEndActuals"]

    for val in range(4, 16):
        customSheet.cell(val, 2).value = round(lastFullYearActuals[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 3).value = round(budget[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 4).value = round(toEndActualsCoppy[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 5).value = round(toEndBp[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 6).value = round(toEndActuals[val - 4] / 10000, 2)
    wb.save(f"{pu}.xlsx")


if __name__ == "__main__":
    make_excel("JAN22", "PU32")
